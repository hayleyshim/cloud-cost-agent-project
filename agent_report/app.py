import redis
import threading
import json
from flask import Flask
from openpyxl import Workbook

app = Flask(__name__)
r = redis.Redis(host='redis', port=6379, db=0)
p = r.pubsub()

def listen_for_messages():
    p.subscribe('cost_channel')
    print("Listening for messages on 'cost_channel'...")
    
    for message in p.listen():
        print("Message received from Redis. Processing...")
        if message['type'] == 'message':
            received_payload = json.loads(message['data'].decode('utf-8'))
            
            # Create a new Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Cost Report"

            # Write a title
            ws['A1'] = 'Cloud Cost Estimation Report'

            # Add headers for the table
            ws['A3'] = 'Service'
            ws['B3'] = 'Estimated Cost (USD)'

            # Write data to the table
            row = 4
            for service, cost in received_payload['cost_breakdown'].items():
                ws[f'A{row}'] = service
                ws[f'B{row}'] = f'${cost:.2f}'
                row += 1

            # Write the total cost
            total_cost_row = row + 1
            ws[f'A{total_cost_row}'] = 'Total Estimated Monthly Cost'
            ws[f'B{total_cost_row}'] = f'${received_payload["total_cost"]:.2f}'

            # Save the Excel file to the correct path
            wb.save('/app/output/cost_report.xlsx')
            print("Cost report saved to /app/output/cost_report.xlsx!")

threading.Thread(target=listen_for_messages, daemon=True).start()

@app.route('/')
def status():
    return 'agent_report is running and listening for cost data!'

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)